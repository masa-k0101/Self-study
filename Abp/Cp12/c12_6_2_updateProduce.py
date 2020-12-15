# -*- coding: utf-8 -*-

#! python3
# updateProduce.py - 農産物スプレッドシートの価格を訂正する

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# 農産物の種類と、更新する価格
PRICE_UPDATE = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

# 行をループして価格を更新する
for row_num in range(2, sheet.max_row + 1):
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATE:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATE[produce_name]

wb.save('updateProduceSales.xlsx')