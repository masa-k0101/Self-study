#! python3
# -*- coding: utf-8 -*-

import openpyxl, smtplib, sys

# スプレッドシートを開き、最近の支払い状況を取得
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

last_col = sheet.get_highest_column()
latest_month = sheet.cell(row=1, column=last_col).value

# 会員の支払い状況を調べる
unpaid_members = {}
for r in range(2, sheet.get_highest_roq() + 1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email

# メールアカウントにログインする
smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login('my_email_address@gmail.com', sys.argv[1])

# リマインダーメールを送信する
for name, email in unpaid_members.items():
    body = """Subject: {} dues unpaid.
Dear {},
Record show that you have not paid dues for {}. Please make this payment
as soon as possible. Thank you!
""".format(latest_month, name, latest_month)
    print('メール送信中 {}...',format(email))
    sendmail_status = smtp_obj.sendmail('my_email_address@gmail.com',email, body)

    if sendmail_status != {}:
        print('{}へメール送信中に問題が起こりました：{}'.format(email, sendmail_status))

smtp_obj.quit()