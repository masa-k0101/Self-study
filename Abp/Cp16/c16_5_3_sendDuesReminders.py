#! python3
# -*- coding: utf-8 -*-

import openpyxl, smtplib, sys

# スプレッドシートを開き、最近の支払い状況を取得
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

last_col = sheet.get_highest_column()
latest_month = sheet.cell(row=1, column=last_col).value

# 会員の支払い状況を調べる
unpaid_members = {}
for r in range(2, sheet.get_highest_roq() + 1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email

# メールアカウントにログインする
smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login('my_email_address@gmail.com', sys.argv[1])

# TODO: リマインダーメールを送信する